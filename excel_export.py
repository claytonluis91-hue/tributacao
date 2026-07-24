"""Exportação profissional e auditável da simulação para Excel."""

import io
from datetime import datetime
from typing import Dict, Iterable, List

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.comments import Comment
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo

from engine_tributario import ConfigTributaria, DadosMes, ResultadoMes


AZUL_ESCURO = "0F172A"
AZUL = "2563EB"
VERDE = "0F766E"
VERDE_CLARO = "CCFBF1"
LARANJA = "F59E0B"
VERMELHO = "DC2626"
CINZA = "E2E8F0"
CINZA_CLARO = "F8FAFC"
BRANCO = "FFFFFF"
TEXTO = "1E293B"
MOEDA = '"R$" #,##0.00'
PERCENTUAL = "0.00%"
INTEIRO = "#,##0"

FONTES = [
    (
        "Simples Nacional: alíquota efetiva, repartição e limite de 5% do ISS",
        "https://normas.receita.fazenda.gov.br/sijut2consulta/link.action?idAto=92278",
    ),
    (
        "Lei Complementar nº 123/2006 e Anexos III, IV e V",
        "https://www.planalto.gov.br/ccivil_03/leis/lcp/lcp123.htm",
    ),
    (
        "IRPJ: 15% e adicional de 10% acima de R$ 20 mil por mês do período",
        "https://www.gov.br/receitafederal/pt-br/assuntos/orientacao-tributaria/tributos/IRPJ",
    ),
    (
        "LC 224/2025: acréscimo de 10% na presunção sobre receita excedente",
        "https://www.gov.br/receitafederal/pt-br/centrais-de-conteudo/publicacoes/perguntas-e-respostas/beneficios-fiscais/perguntas-e-respostas-reducao-dos-incentivos-e-beneficios-tributarios-v3-final.pdf",
    ),
    (
        "Reforma Tributária do Consumo: orientações para 2026",
        "https://www.gov.br/receitafederal/pt-br/acesso-a-informacao/acoes-e-programas/programas-e-atividades/reforma-tributaria-do-consumo/orientacoes-2026",
    ),
]


def _titulo(ws, titulo: str, subtitulo: str, ultima_coluna: int) -> None:
    ws.sheet_view.showGridLines = False
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ultima_coluna)
    ws["A1"] = titulo
    ws["A1"].font = Font(name="Aptos Display", size=20, bold=True, color=BRANCO)
    ws["A1"].fill = PatternFill("solid", fgColor=AZUL_ESCURO)
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 34
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ultima_coluna)
    ws["A2"] = subtitulo
    ws["A2"].font = Font(name="Aptos", size=10, color="475569")
    ws["A2"].fill = PatternFill("solid", fgColor=CINZA_CLARO)
    ws["A2"].alignment = Alignment(vertical="center")
    ws.row_dimensions[2].height = 24


def _cabecalho(ws, linha: int, colunas: int) -> None:
    faixa = ws.cell(linha, 1).parent.iter_rows(
        min_row=linha, max_row=linha, min_col=1, max_col=colunas
    )
    for row in faixa:
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=VERDE)
            cell.font = Font(name="Aptos", size=10, bold=True, color=BRANCO)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[linha].height = 32


def _tabela(ws, nome: str, referencia: str) -> None:
    tabela = Table(displayName=nome, ref=referencia)
    tabela.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tabela)


def _borda_card():
    lado = Side(style="thin", color=CINZA)
    return Border(left=lado, right=lado, top=lado, bottom=lado)


def _ajustar_larguras(ws, larguras: Dict[str, float]) -> None:
    for coluna, largura in larguras.items():
        ws.column_dimensions[coluna].width = largura


def _melhor_regime(resultado: ResultadoMes):
    custos = {
        "Simples Nacional": resultado.total_sn,
        "Lucro Presumido": resultado.total_lp,
        "Lucro Real": resultado.total_lr,
    }
    ordenados = sorted(custos.items(), key=lambda item: item[1])
    return ordenados[0][0], ordenados[1][1] - ordenados[0][1]


def _painel(wb: Workbook, resultados: List[ResultadoMes], metadados: Dict) -> None:
    ws = wb.active
    ws.title = "Painel"
    _titulo(
        ws,
        "Simulação tributária | Painel executivo",
        f"Visão comparativa de 12 meses • gerado em {datetime.now():%d/%m/%Y %H:%M}",
        12,
    )
    totais = {
        "Simples Nacional": sum(r.total_sn for r in resultados),
        "Lucro Presumido": sum(r.total_lp for r in resultados),
        "Lucro Real": sum(r.total_lr for r in resultados),
    }
    faturamento = sum(r.faturamento for r in resultados)
    ranking = sorted(totais.items(), key=lambda item: item[1])
    economia = ranking[1][1] - ranking[0][1]

    cards = [
        ("A4:C4", "A5:C6", "REGIME MAIS ECONÔMICO", ranking[0][0], VERDE_CLARO),
        ("D4:F4", "D5:F6", "ECONOMIA VS. 2º COLOCADO", economia, "DBEAFE"),
        ("G4:I4", "G5:I6", "FATURAMENTO PROJETADO", faturamento, "FEF3C7"),
        (
            "J4:L4",
            "J5:L6",
            "ALÍQUOTA EFETIVA VENCEDORA",
            ranking[0][1] / faturamento if faturamento else 0,
            "FEE2E2",
        ),
    ]
    for titulo_ref, valor_ref, titulo, valor, cor in cards:
        ws.merge_cells(titulo_ref)
        ws.merge_cells(valor_ref)
        titulo_cell = ws[titulo_ref.split(":")[0]]
        valor_cell = ws[valor_ref.split(":")[0]]
        titulo_cell.value = titulo
        titulo_cell.font = Font(size=9, bold=True, color="475569")
        titulo_cell.alignment = Alignment(horizontal="center", vertical="center")
        titulo_cell.fill = PatternFill("solid", fgColor=cor)
        valor_cell.value = valor
        valor_cell.font = Font(size=16, bold=True, color=AZUL_ESCURO)
        valor_cell.alignment = Alignment(horizontal="center", vertical="center")
        valor_cell.fill = PatternFill("solid", fgColor=cor)
        if isinstance(valor, float):
            valor_cell.number_format = PERCENTUAL if "ALÍQUOTA" in titulo else MOEDA
        for row in ws[titulo_ref.split(":")[0] : valor_ref.split(":")[1]]:
            for cell in row:
                cell.border = _borda_card()

    ws["A8"] = "Regime"
    ws["B8"] = "Total projetado"
    ws["C8"] = "Alíquota efetiva"
    ws["D8"] = "Diferença para o menor"
    for idx, (regime, total) in enumerate(totais.items(), 9):
        ws.cell(idx, 1, regime)
        ws.cell(idx, 2, total)
        ws.cell(idx, 3, total / faturamento if faturamento else 0)
        ws.cell(idx, 4, total - ranking[0][1])
    _cabecalho(ws, 8, 4)
    for row in ws.iter_rows(min_row=9, max_row=11, min_col=2, max_col=4):
        row[0].number_format = MOEDA
        row[1].number_format = PERCENTUAL
        row[2].number_format = MOEDA
    _tabela(ws, "ResumoRegimes", "A8:D11")

    chart = LineChart()
    chart.title = "Carga tributária mês a mês"
    chart.y_axis.title = "R$"
    chart.x_axis.title = "Competência"
    chart.style = 13
    chart.height = 8.5
    chart.width = 16.5
    mensal = wb["Apuração Mensal"]
    dados = Reference(mensal, min_col=9, max_col=11, min_row=4, max_row=4 + len(resultados))
    categorias = Reference(mensal, min_col=1, min_row=5, max_row=4 + len(resultados))
    chart.add_data(dados, titles_from_data=True)
    chart.set_categories(categorias)
    for serie, titulo in zip(
        chart.series, ("Simples Nacional", "Lucro Presumido", "Lucro Real")
    ):
        serie.tx = SeriesLabel(v=titulo)
    chart.legend.position = "b"
    ws.add_chart(chart, "F8")

    ws.merge_cells("A23:L23")
    ws["A23"] = "Leitura rápida"
    ws["A23"].fill = PatternFill("solid", fgColor=AZUL_ESCURO)
    ws["A23"].font = Font(bold=True, color=BRANCO)
    ws.merge_cells("A24:L26")
    ws["A24"] = (
        "A comparação é uma simulação gerencial. IRPJ e CSLL do Lucro Presumido e do "
        "Lucro Real são fechados por trimestre; o adicional aparece no último mês do "
        "trimestre. Benefícios, retenções, compensações, particularidades municipais e "
        "limitações de crédito devem ser validados pelo responsável tributário."
    )
    ws["A24"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["A24"].fill = PatternFill("solid", fgColor=CINZA_CLARO)
    ws["A24"].border = _borda_card()
    ws["A28"] = "Arquivo-base"
    ws["B28"] = metadados.get("arquivo_pgdas", "Entrada manual")
    ws["A29"] = "Período-base"
    ws["B29"] = metadados.get("pa", "Não informado")
    _ajustar_larguras(
        ws,
        {col: 14 for col in "ABCDEFGHIJKL"}
        | {"A": 22, "B": 21, "C": 18, "D": 22, "F": 16},
    )
    ws.freeze_panes = "A4"


def _apuracao_mensal(wb: Workbook, resultados: List[ResultadoMes]) -> None:
    ws = wb.create_sheet("Apuração Mensal")
    _titulo(
        ws,
        "Apuração simulada mês a mês",
        "RBT12 e FS12 usam a janela móvel dos 12 meses anteriores a cada competência.",
        16,
    )
    cabecalhos = [
        "Competência",
        "Faturamento",
        "Folha",
        "RBT12",
        "FS12",
        "Fator R",
        "Anexo",
        "Alíquota efetiva SN",
        "Simples Nacional",
        "Lucro Presumido",
        "Lucro Real",
        "Melhor regime",
        "Economia para o 2º",
        "SN / Receita",
        "LP / Receita",
        "LR / Receita",
    ]
    ws.append([])
    ws.append(cabecalhos)
    for resultado in resultados:
        melhor, economia = _melhor_regime(resultado)
        ws.append(
            [
                resultado.competencia,
                resultado.faturamento,
                resultado.folha,
                resultado.sn_rbt12,
                resultado.sn_fs12,
                resultado.sn_fator_r,
                resultado.sn_anexo,
                resultado.sn_aliquota_efetiva,
                resultado.total_sn,
                resultado.total_lp,
                resultado.total_lr,
                melhor,
                economia,
                resultado.total_sn / resultado.faturamento if resultado.faturamento else 0,
                resultado.total_lp / resultado.faturamento if resultado.faturamento else 0,
                resultado.total_lr / resultado.faturamento if resultado.faturamento else 0,
            ]
        )
    _cabecalho(ws, 4, len(cabecalhos))
    final = 4 + len(resultados)
    _tabela(ws, "ApuracaoMensal", f"A4:P{final}")
    for row in ws.iter_rows(min_row=5, max_row=final):
        for col in (2, 3, 4, 5, 9, 10, 11, 13):
            row[col - 1].number_format = MOEDA
        for col in (6, 8, 14, 15, 16):
            row[col - 1].number_format = PERCENTUAL
        row[6].alignment = Alignment(horizontal="center")
    ws.conditional_formatting.add(
        f"M5:M{final}",
        ColorScaleRule(
            start_type="min",
            start_color="DCFCE7",
            mid_type="percentile",
            mid_value=50,
            mid_color="FEF3C7",
            end_type="max",
            end_color="FEE2E2",
        ),
    )
    _ajustar_larguras(
        ws,
        {
            "A": 14,
            "B": 16,
            "C": 15,
            "D": 16,
            "E": 16,
            "F": 12,
            "G": 10,
            "H": 16,
            "I": 18,
            "J": 18,
            "K": 16,
            "L": 19,
            "M": 19,
            "N": 13,
            "O": 13,
            "P": 13,
        },
    )
    ws.freeze_panes = "B5"
    ws.auto_filter.ref = f"A4:P{final}"


def _simples(wb: Workbook, resultados: List[ResultadoMes]) -> None:
    ws = wb.create_sheet("Simples Nacional")
    _titulo(
        ws,
        "Simples Nacional | Memória de cálculo",
        "DAS segregado por tributo. No Anexo IV, a contribuição patronal é calculada fora do DAS.",
        19,
    )
    headers = [
        "Competência",
        "Faturamento",
        "RBT12",
        "FS12",
        "Fator R",
        "Anexo",
        "Faixa",
        "Alíquota nominal",
        "Alíquota efetiva",
        "IRPJ",
        "CSLL",
        "Cofins",
        "PIS/Pasep",
        "CPP no DAS",
        "ISS",
        "INSS fora do DAS",
        "Total DAS",
        "Total do regime",
        "Alíquota total",
    ]
    ws.append([])
    ws.append(headers)
    for idx, r in enumerate(resultados, 5):
        inss_extra = max(r.sn_encargos - r.sn_valor_cpp, 0.0)
        ws.append(
            [
                r.competencia,
                r.faturamento,
                r.sn_rbt12,
                r.sn_fs12,
                r.sn_fator_r,
                r.sn_anexo,
                r.sn_faixa,
                r.sn_aliquota_nominal,
                r.sn_aliquota_efetiva,
                r.sn_valor_irpj,
                r.sn_valor_csll,
                r.sn_valor_cofins,
                r.sn_valor_pis,
                r.sn_valor_cpp,
                r.sn_valor_iss,
                inss_extra,
                f"=SUM(J{idx}:O{idx})",
                f"=Q{idx}+P{idx}",
                f'=IFERROR(R{idx}/B{idx},0)',
            ]
        )
    _cabecalho(ws, 4, len(headers))
    final = 4 + len(resultados)
    _tabela(ws, "MemoriaSimples", f"A4:S{final}")
    for row in ws.iter_rows(min_row=5, max_row=final):
        for col in (2, 3, 4, 10, 11, 12, 13, 14, 15, 16, 17, 18):
            row[col - 1].number_format = MOEDA
        for col in (5, 8, 9, 19):
            row[col - 1].number_format = PERCENTUAL
        row[5].alignment = Alignment(horizontal="center")
    _ajustar_larguras(ws, {chr(64 + i): 15 for i in range(1, 20)} | {"A": 14, "F": 9})
    ws.freeze_panes = "B5"


def _presumido(wb: Workbook, resultados: List[ResultadoMes]) -> None:
    ws = wb.create_sheet("Lucro Presumido")
    _titulo(
        ws,
        "Lucro Presumido | Memória de cálculo",
        "IRPJ/CSLL trimestrais; PIS, Cofins, ISS e encargos apresentados por competência.",
        14,
    )
    headers = [
        "Competência",
        "Faturamento",
        "Trimestre",
        "Base IRPJ",
        "Base CSLL",
        "IRPJ 15%",
        "CSLL 9%",
        "Adicional IRPJ 10%",
        "PIS 0,65%",
        "Cofins 3%",
        "ISS",
        "INSS patronal",
        "Total do regime",
        "Alíquota efetiva",
    ]
    ws.append([])
    ws.append(headers)
    for idx, r in enumerate(resultados, 5):
        ws.append(
            [
                r.competencia,
                r.faturamento,
                r.lp_trimestre,
                r.lp_base_irpj,
                r.lp_base_csll,
                r.lp_irpj,
                r.lp_csll,
                r.lp_adicional_irpj,
                r.lp_pis,
                r.lp_cofins,
                r.lp_valor_iss,
                r.lp_encargos,
                f"=SUM(F{idx}:L{idx})",
                f'=IFERROR(M{idx}/B{idx},0)',
            ]
        )
    _cabecalho(ws, 4, len(headers))
    final = 4 + len(resultados)
    _tabela(ws, "MemoriaPresumido", f"A4:N{final}")
    for row in ws.iter_rows(min_row=5, max_row=final):
        for col in range(2, 14):
            if col != 3:
                row[col - 1].number_format = MOEDA
        row[13].number_format = PERCENTUAL
    _ajustar_larguras(ws, {chr(64 + i): 16 for i in range(1, 15)} | {"A": 14, "C": 12})
    ws.freeze_panes = "B5"


def _real(wb: Workbook, resultados: List[ResultadoMes]) -> None:
    ws = wb.create_sheet("Lucro Real")
    _titulo(
        ws,
        "Lucro Real trimestral | Memória de cálculo",
        "Créditos de PIS/Cofins limitados aos débitos informados; validar a elegibilidade documental.",
        16,
    )
    headers = [
        "Competência",
        "Faturamento",
        "Folha",
        "Despesas dedutíveis",
        "Base de créditos",
        "Trimestre",
        "PIS débito",
        "Cofins débito",
        "Créditos",
        "ISS",
        "INSS patronal",
        "Lucro antes IR/CSLL",
        "IRPJ 15%",
        "CSLL 9%",
        "Adicional IRPJ",
        "Total do regime",
    ]
    ws.append([])
    ws.append(headers)
    for idx, r in enumerate(resultados, 5):
        ws.append(
            [
                r.competencia,
                r.faturamento,
                r.folha,
                r.despesas_dedutiveis,
                r.compras_credito,
                r.lr_trimestre,
                r.lr_pis_debito,
                r.lr_cofins_debito,
                r.lr_creditos,
                r.lr_valor_iss,
                r.lr_encargos,
                r.lr_lair,
                r.lr_irpj,
                r.lr_csll,
                r.lr_adicional_irpj,
                f"=SUM(G{idx}:H{idx})-I{idx}+SUM(J{idx}:K{idx})+SUM(M{idx}:O{idx})",
            ]
        )
    _cabecalho(ws, 4, len(headers))
    final = 4 + len(resultados)
    _tabela(ws, "MemoriaReal", f"A4:P{final}")
    for row in ws.iter_rows(min_row=5, max_row=final):
        for col in range(2, 17):
            if col != 6:
                row[col - 1].number_format = MOEDA
    _ajustar_larguras(ws, {chr(64 + i): 16 for i in range(1, 17)} | {"A": 14, "D": 19, "F": 12})
    ws.freeze_panes = "B5"


def _historico(wb: Workbook, historico: Iterable[Dict]) -> None:
    ws = wb.create_sheet("Histórico")
    _titulo(
        ws,
        "Histórico usado na simulação",
        "As 12 competências abaixo formam a primeira janela móvel de RBT12 e FS12.",
        4,
    )
    ws.append([])
    ws.append(["Competência", "Faturamento", "Folha", "Origem"])
    linhas = list(historico)
    for item in linhas:
        ws.append(
            [
                item.get("Competência", ""),
                float(item.get("Faturamento", 0)),
                float(item.get("Folha", 0)),
                item.get("Origem", "Entrada manual"),
            ]
        )
    _cabecalho(ws, 4, 4)
    if linhas:
        _tabela(ws, "HistoricoBase", f"A4:D{4 + len(linhas)}")
    for row in ws.iter_rows(min_row=5, max_row=4 + len(linhas), min_col=2, max_col=3):
        for cell in row:
            cell.number_format = MOEDA
    _ajustar_larguras(ws, {"A": 15, "B": 20, "C": 20, "D": 28})
    ws.freeze_panes = "A5"


def _premissas(wb: Workbook, config: ConfigTributaria) -> None:
    ws = wb.create_sheet("Premissas e Fontes")
    _titulo(
        ws,
        "Premissas, escopo e fontes",
        "Parâmetros editados na aplicação e referências oficiais consultadas.",
        5,
    )
    ws["A4"] = "Premissa"
    ws["B4"] = "Valor"
    ws["C4"] = "Observação"
    premissas = [
        ("Ano de referência", config.ano_referencia, "Regras transitórias vigentes no ano."),
        ("RAT", config.aliquota_rat, "Aplicado com o FAP sobre a folha."),
        ("FAP", config.fap, "Multiplicador do RAT."),
        ("Terceiros", config.aliquota_terceiros, "Parâmetro variável conforme enquadramento."),
        ("ISS", config.aliquota_issqn, "Confirmar a alíquota e o município competente."),
        ("Presunção IRPJ", config.presuncao_irpj, "Percentual-base informado."),
        ("Presunção CSLL", config.presuncao_csll, "Percentual-base informado."),
        ("Atividade no Anexo IV", "Sim" if config.is_anexo_iv else "Não", "CPP fora do DAS no Anexo IV."),
        (
            "Atividade sujeita ao Fator R",
            "Sim" if config.fator_r_sujeito else "Não",
            "Anexo III se Fator R ≥ 28%; Anexo V abaixo de 28%.",
        ),
    ]
    for item in premissas:
        ws.append(item)
    _cabecalho(ws, 4, 3)
    for linha in (6, 8, 9, 10, 11):
        ws.cell(linha, 2).number_format = PERCENTUAL
    ws["A15"] = "Referência"
    ws["B15"] = "URL oficial"
    ws["C15"] = "Uso no modelo"
    _cabecalho(ws, 15, 3)
    for descricao, url in FONTES:
        ws.append([descricao, url, "Base normativa / orientação"])
        ws.cell(ws.max_row, 2).hyperlink = url
        ws.cell(ws.max_row, 2).style = "Hyperlink"
        for cell in ws[ws.max_row][:3]:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[ws.max_row].height = 70
    ws["A23"] = "Limitações relevantes"
    ws["A23"].font = Font(bold=True, color=BRANCO)
    ws["A23"].fill = PatternFill("solid", fgColor=AZUL_ESCURO)
    ws.merge_cells("A23:E23")
    ws.merge_cells("A24:E27")
    ws["A24"] = (
        "Modelo comparativo, não substitui apuração fiscal. Não contempla retenções, benefícios, "
        "receitas monofásicas, substituição tributária, exportações, prejuízos fiscais de períodos "
        "anteriores, adicional estadual de ICMS, regras setoriais, limites individualizados de "
        "crédito nem obrigações acessórias. Em 2026, IBS/CBS estão no ano-teste e a dispensa de "
        "recolhimento depende do cumprimento das obrigações aplicáveis."
    )
    ws["A24"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["A24"].fill = PatternFill("solid", fgColor=CINZA_CLARO)
    ws["A24"].border = _borda_card()
    ws.row_dimensions[24].height = 78
    ws["B5"].comment = Comment(
        "Parâmetros devem ser confirmados para o CNPJ, atividade e município.", "User"
    )
    _ajustar_larguras(ws, {"A": 55, "B": 60, "C": 28, "D": 14, "E": 14})
    ws.freeze_panes = "A4"


def gerar_excel(
    resultados: List[ResultadoMes],
    dados_input: List[DadosMes],
    config: ConfigTributaria,
    historico: Iterable[Dict],
    metadados: Dict | None = None,
) -> bytes:
    del dados_input  # resultados já mantêm os dados de entrada necessários
    wb = Workbook()
    # A planilha mensal precisa existir antes do gráfico do painel.
    wb.active.title = "Painel"
    _apuracao_mensal(wb, resultados)
    wb.active = 0
    _painel(wb, resultados, metadados or {})
    _simples(wb, resultados)
    _presumido(wb, resultados)
    _real(wb, resultados)
    _historico(wb, historico)
    _premissas(wb, config)
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
