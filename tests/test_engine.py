import io
import unittest
from pathlib import Path

from openpyxl import load_workbook

from engine_tributario import ConfigTributaria, DadosMes, simular_12_meses
from excel_export import gerar_excel
from pgdas_parser import extrair_dados_pgdas


class MotorTributarioTest(unittest.TestCase):
    def test_reconcilia_pgdas_anexo_v(self):
        arquivos = list((Path(__file__).parents[1] / "documentos").glob("*.pdf"))
        if not arquivos:
            self.skipTest("Extrato PGDAS não disponível")
        extraido = extrair_dados_pgdas(arquivos[0].read_bytes())
        # Valores mensais que formam o RBT12 declarado do PA 05/2026.
        historico = extraido["historico"]
        receitas = [item["Faturamento"] for item in historico[:-1]]
        # O parser usa 11 anteriores + PA; para reconciliar o PA, recupera o
        # primeiro mês da janela declarada a partir do total do extrato.
        receitas.insert(0, extraido["rbt12"] - sum(receitas))
        receitas = receitas[:12]
        folhas = [extraido["folha12"] / 12] * 12
        resultado = simular_12_meses(
            [DadosMes(1, extraido["rpa"], 0, 0, 0, extraido["pa"])],
            ConfigTributaria(),
            historico_faturamento=receitas,
            historico_folha=folhas,
        )[0]
        self.assertEqual(resultado.sn_anexo, "V")
        self.assertEqual(resultado.sn_faixa, 4)
        self.assertAlmostEqual(resultado.sn_total_das, 30_914.32, places=2)
        self.assertAlmostEqual(resultado.sn_valor_irpj, 6_492.01, places=2)
        self.assertAlmostEqual(resultado.sn_valor_iss, 6_492.01, places=2)

    def test_rbt12_remove_mes_real_mais_antigo(self):
        historico = [10_000 + i * 1_000 for i in range(12)]
        dados = [
            DadosMes(1, 50_000, 10_000, 0, 0, "01/2026"),
            DadosMes(2, 60_000, 10_000, 0, 0, "02/2026"),
        ]
        resultados = simular_12_meses(
            dados,
            ConfigTributaria(),
            historico_faturamento=historico,
            historico_folha=[10_000] * 12,
        )
        self.assertEqual(resultados[0].sn_rbt12, sum(historico))
        self.assertEqual(resultados[1].sn_rbt12, sum(historico[1:]) + 50_000)

    def test_anexo_iv_nao_inclui_cpp_no_das(self):
        resultado = simular_12_meses(
            [DadosMes(1, 100_000, 20_000, 0, 0, "01/2026")],
            ConfigTributaria(is_anexo_iv=True),
            historico_faturamento=[100_000] * 12,
            historico_folha=[20_000] * 12,
        )[0]
        self.assertEqual(resultado.sn_anexo, "IV")
        self.assertEqual(resultado.sn_valor_cpp, 0)
        self.assertGreater(resultado.sn_encargos, 0)

    def test_excel_contem_abas_formulas_e_fontes(self):
        historico = [
            {
                "Competência": f"{mes:02d}/2025",
                "Faturamento": 100_000,
                "Folha": 30_000,
                "Origem": "Teste",
            }
            for mes in range(1, 13)
        ]
        dados = [
            DadosMes(mes, 110_000, 32_000, 55_000, 10_000, f"{mes:02d}/2026")
            for mes in range(1, 13)
        ]
        config = ConfigTributaria(aliquota_issqn=0.02)
        resultados = simular_12_meses(
            dados,
            config,
            historico_faturamento=[100_000] * 12,
            historico_folha=[30_000] * 12,
        )
        arquivo = gerar_excel(resultados, dados, config, historico)
        wb = load_workbook(io.BytesIO(arquivo), data_only=False)
        self.assertEqual(
            wb.sheetnames,
            [
                "Painel",
                "Apuração Mensal",
                "Simples Nacional",
                "Lucro Presumido",
                "Lucro Real",
                "Histórico",
                "Premissas e Fontes",
            ],
        )
        self.assertTrue(str(wb["Simples Nacional"]["Q5"].value).startswith("="))
        self.assertIn("receita.fazenda.gov.br", wb["Premissas e Fontes"]["B16"].value)


if __name__ == "__main__":
    unittest.main()
